import os
import shutil
import time
import openpyxl

def main():
    root_list = []
    root_temp = []
    time_list = []
    del_dirs_list = []
    for root, dirs, files in os.walk(r'User'):
        # print("當前目錄路徑:",root) #当前目录路径
        # print("當前路徑下所有子目錄:",dirs) #当前路径下所有子目录
        # print("當前路徑下所有非目錄子文件:",files) #当前路径下所有非目录子文件
        root_list.append(root)
        # print("root_list:",root_list)
        # print("dir_list:",dirs) #dirs&files本身就是list

        for d in range(len(dirs)):
            # print("dirs[d]:",dirs[d])
            if dirs[d] == "windows":
                root_temp.append(root)
                # print("root_temp:",root_temp)
                for r in range(len(root_temp)):
                    # print("root_temp[r]:",root_temp[r])
                    del_dirs = root_temp[r] + '\\' + dirs[d]
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                # print(del_dirs)
                # print(current_time)
                shutil.rmtree(del_dirs)
                time_list.append(current_time)
                del_dirs_list.append(del_dirs)
                # print(time_list)
                # print(del_dirs_list)
    write_excel_xlsx(time_list, del_dirs_list, file_name_xlsx, sheet_name_xlsx) #file_name_xlsx, sheet_name_xlsx定義在下面
    return time_list, del_dirs_list

def write_excel_xlsx(time_list, del_dirs_list, path, sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["時間","執行","結果","路徑"])
    for i in range(0, len(time_list)):
        # print(time_list[i])
        sheet.cell(row=i+2, column=1, value=time_list[i])
        sheet.cell(row=i+2, column=2, value="delete")
        sheet.cell(row=i+2, column=3, value="successfully deleted")
    for i in range(0, len(del_dirs_list)):
        # print(del_dirs_list[i])
        sheet.cell(row=i+2, column=4, value=del_dirs_list[i])
    workbook.save(path)
    print("xlsx格式表格寫入成功！")
    
file_name_xlsx = "Log\\delete_record.xlsx"
sheet_name_xlsx = "刪除檔案紀錄表"

main()